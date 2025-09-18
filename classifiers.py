import numpy as np
import pandas as pd
from openpyxl import load_workbook
from sklearn import linear_model
from sklearn.model_selection import train_test_split
from time import time, process_time

# load and summarize the dataset
from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error

# to write in excel
import openpyxl

# for training dataset
from sklearn.preprocessing import MinMaxScaler

dataframe = pd.read_excel(r'D:\Thesis\Kidney_Disease\Dataset\dataset.xlsx')
#print("Full dataset: ", dataframe)

# imputer = KNNImputer(n_neighbors=2, weights="uniform")
# print(imputer.fit_transform(dataframe))
# print("Imputed dataset: ", dataframe)


# load the dataset
data_training = dataframe.values
X, y = data_training[:, :-1], data_training[:, -1]

# Normalization the dataset
scaler = MinMaxScaler()

fit_X = scaler.fit(X)
X = scaler.transform(X)

X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.25, random_state=42)

total_testingData = len(X_test)


# Ridge Classifier
# .............................

# training the model
ridge = linear_model.RidgeClassifier(alpha=1)

# Fit the model using training dataset
ridge.fit(X_train, y_train)

# calculating score
print('Training Score = ', ridge.score(X_train, y_train))

# calculating disease prediction
disease_predict = ridge.predict(X_test)

# writting actual and predicted result in excel
excel_file = load_workbook(r'D:\Thesis\Kidney_Disease\Dataset\score.xlsx')

excel_sheet = excel_file.create_sheet("Sheet1")
excel_sheet.title = "Ridge"

excel_sheet['A1'] = 'Index'
excel_sheet['B1'] = 'Actual Result'
excel_sheet['C1'] = 'Predicted Result'
excel_sheet['D1'] = 'Result'
excel_sheet['E1'] = 'Score in %'

row = 2
col = 1

for index in range(1, total_testingData+1):
     excel_sheet.cell(row = row, column = col).value = index
     row += 1

row = 2
col = 2

error_count = 0
for index1, index2 in zip(y_test, disease_predict):
     excel_sheet.cell(row=row, column=col).value = index1
     excel_sheet.cell(row=row, column=col+1).value = index2
     if(index1 == index2):
          excel_sheet.cell(row=row, column=col+2).value = "Yes"
     else:
          excel_sheet.cell(row=row, column=col+2).value = "No"
          excel_sheet.cell(row=row, column=col+2).font = openpyxl.styles.Font(color='00FF0000')
          error_count += 1
     row += 1

# calculating score
score = ((total_testingData - error_count)*100)/total_testingData
print('Testing Score = ', score)


excel_sheet['E2'] = score

# Support Vector Classifier
# .............................

from sklearn.svm import SVC

# training the model
svc = SVC(kernel='linear')

# Fit the model using training dataset
svc.fit(X_train, y_train)

# calculating score
print('Training Score = ', svc.score(X_train, y_train))

# calculating disease prediction
disease_predict = svc.predict(X_test)

excel_sheet = excel_file.create_sheet("Sheet1")
excel_sheet.title = "SVC"



excel_sheet['A1'] = 'Index'
excel_sheet['B1'] = 'Actual Result'
excel_sheet['C1'] = 'Predicted Result'
excel_sheet['D1'] = 'Result'
excel_sheet['E1'] = 'Score in %'

row = 2
col = 1

for index in range(1, total_testingData+1):
     excel_sheet.cell(row = row, column = col).value = index
     row += 1

row = 2
col = 2

error_count = 0
for index1, index2 in zip(y_test, disease_predict):
     excel_sheet.cell(row=row, column=col).value = index1
     excel_sheet.cell(row=row, column=col+1).value = index2
     if(index1 == index2):
          excel_sheet.cell(row=row, column=col+2).value = "Yes"
     else:
          excel_sheet.cell(row=row, column=col+2).value = "No"
          excel_sheet.cell(row=row, column=col+2).font = openpyxl.styles.Font(color='00FF0000')
          error_count += 1
     row += 1

# calculating score
score = ((total_testingData - error_count)*100)/total_testingData
print('Testing Score = ', score)

excel_sheet['E2'] = score


# Decision Tree Classifiers
# .............................

from sklearn import tree

# training the model
dt = tree.DecisionTreeClassifier()

# Fit the model using training dataset
dt.fit(X_train, y_train)

# calculating score
print('Training Score = ', dt.score(X_train, y_train))

# calculating disease prediction
disease_predict = dt.predict(X_test)

excel_sheet = excel_file.create_sheet("Sheet1")
excel_sheet.title = "DT"



excel_sheet['A1'] = 'Index'
excel_sheet['B1'] = 'Actual Result'
excel_sheet['C1'] = 'Predicted Result'
excel_sheet['D1'] = 'Result'
excel_sheet['E1'] = 'Score in %'

row = 2
col = 1

for index in range(1, total_testingData+1):
     excel_sheet.cell(row = row, column = col).value = index
     row += 1

row = 2
col = 2

error_count = 0
for index1, index2 in zip(y_test, disease_predict):
     excel_sheet.cell(row=row, column=col).value = index1
     excel_sheet.cell(row=row, column=col+1).value = index2
     if(index1 == index2):
          excel_sheet.cell(row=row, column=col+2).value = "Yes"
     else:
          excel_sheet.cell(row=row, column=col+2).value = "No"
          excel_sheet.cell(row=row, column=col+2).font = openpyxl.styles.Font(color='00FF0000')
          error_count += 1
     row += 1

# calculating score
score = ((total_testingData - error_count)*100)/total_testingData
print('Testing Score = ', score)

excel_sheet['E2'] = score


# Decision Tree Classifiers
# .............................

from sklearn.ensemble import RandomForestClassifier

# training the model
randomForest = RandomForestClassifier(n_estimators=50)

# Fit the model using training dataset
randomForest.fit(X_train, y_train)

# calculating score
print('Training Score = ', randomForest.score(X_train, y_train))

# calculating disease prediction
disease_predict = randomForest.predict(X_test)

excel_sheet = excel_file.create_sheet("Sheet1")
excel_sheet.title = "Random Forest"



excel_sheet['A1'] = 'Index'
excel_sheet['B1'] = 'Actual Result'
excel_sheet['C1'] = 'Predicted Result'
excel_sheet['D1'] = 'Result'
excel_sheet['E1'] = 'Score in %'

row = 2
col = 1

for index in range(1, total_testingData+1):
     excel_sheet.cell(row = row, column = col).value = index
     row += 1

row = 2
col = 2

error_count = 0
for index1, index2 in zip(y_test, disease_predict):
     excel_sheet.cell(row=row, column=col).value = index1
     excel_sheet.cell(row=row, column=col+1).value = index2
     if(index1 == index2):
          excel_sheet.cell(row=row, column=col+2).value = "Yes"
     else:
          excel_sheet.cell(row=row, column=col+2).value = "No"
          excel_sheet.cell(row=row, column=col+2).font = openpyxl.styles.Font(color='00FF0000')
          error_count += 1
     row += 1

# calculating score
score = ((total_testingData - error_count)*100)/total_testingData
print('Testing Score = ', score)

excel_sheet['E2'] = score

excel_file.save(r'D:\Thesis\Kidney_Disease\Dataset\score.xlsx')

excel_file.close()